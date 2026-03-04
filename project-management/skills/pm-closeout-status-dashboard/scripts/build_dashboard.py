#!/usr/bin/env python3
"""
build_dashboard.py — Generate an interactive HTML closeout dashboard from an Excel spreadsheet.

Supports two input formats (auto-detected from headers):
  1. Manual closeout tracking spreadsheet (Subcontractor, Spec, Warranty, O&M, etc.)
  2. Procore submittal log export (Spec Section, Title, Type, Status, Responsible Contractor, etc.)

Usage:
    python3 build_dashboard.py --input "path/to/Closeout Status.xlsx" --output "path/to/Dashboard.html" --title "Project Name — Closeout Dashboard"
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import date

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------

MANUAL_HEADER_PATTERNS = {
    "sub": [r"subcontractor", r"\bsub\b"],
    "spec": [r"\bspec\b"],
    "desc": [r"description", r"\bdesc\b"],
    "warranty": [r"warranty"],
    "om": [r"o\s*&\s*m", r"o\s+and\s+m", r"operation"],
    "asbuilt": [r"as[\s-]?built"],
    "attic": [r"attic"],
    "training": [r"training"],
    "notes": [r"\bnote"],
}

PROCORE_REQUIRED_HEADERS = ["type", "status", "responsible contractor", "title"]


def detect_format(ws):
    """Return 'procore' or 'manual' based on row-1 headers."""
    headers_lower = []
    for cell in ws[1]:
        val = str(cell.value or "").strip().lower()
        headers_lower.append(val)

    # Check for Procore signature headers
    procore_matches = sum(1 for h in PROCORE_REQUIRED_HEADERS if h in headers_lower)
    if procore_matches >= 3:
        return "procore"
    return "manual"


def detect_manual_columns(ws):
    """Return dict mapping category key -> column index (0-based) from row 1 headers."""
    cols = {}
    for idx, cell in enumerate(ws[1]):
        val = str(cell.value or "").strip().lower()
        if not val:
            continue
        for key, patterns in MANUAL_HEADER_PATTERNS.items():
            if key in cols:
                continue
            for pat in patterns:
                if re.search(pat, val):
                    cols[key] = idx
                    break
    if "sub" not in cols:
        print("ERROR: Could not find a Subcontractor column in row 1.")
        sys.exit(1)
    return cols


def detect_procore_columns(ws):
    """Return dict mapping procore field -> column index (0-based) from row 1 headers."""
    cols = {}
    field_patterns = {
        "spec_section": [r"spec\s*section"],
        "number": [r"^#$"],
        "title": [r"^title$"],
        "type": [r"^type$"],
        "description": [r"^description$"],
        "responsible_contractor": [r"responsible\s*contractor"],
        "status": [r"^status$"],
        "response": [r"^response$"],
    }
    for idx, cell in enumerate(ws[1]):
        val = str(cell.value or "").strip().lower()
        if not val:
            continue
        for key, patterns in field_patterns.items():
            if key in cols:
                continue
            for pat in patterns:
                if re.search(pat, val):
                    cols[key] = idx
                    break
    return cols


# ---------------------------------------------------------------------------
# Sub name normalization
# ---------------------------------------------------------------------------

DEFAULT_NAME_MAP = {
    "CR Lighting": "CR Lighting & Electric",
    "Moore's Floores": "Moore's Floors",
    "Ritters Landscape": "Ritters Landscaping",
    "Ritter's Landscaping": "Ritters Landscaping",
    "Continental Fire Sprinkler Company": "Continental Fire Sprinkler",
    "Prestige Coatings": "Prestige Coatings and Color",
    "Sparrow Plumbing & Heating": "Sparrow Plumbing",
    "Mid-States/East Moline": "Mid-States Door & Hardware",
    "East Moline Glass": "East Moline Glass",
    "Fetzers": "Fetzers",
}

SKIP_NAMES = {"as built", "as builts received", "as-built", ""}


def load_name_map(path):
    """Load a JSON file of name mappings and merge with defaults."""
    mapping = dict(DEFAULT_NAME_MAP)
    if path:
        with open(path) as f:
            mapping.update(json.load(f))
    return mapping


def normalize_sub(name, name_map):
    """Normalize sub name: strip suffixes, apply mapping."""
    if not name:
        return None
    n = str(name).strip()
    if n.lower() in SKIP_NAMES:
        return None
    # Strip common suffixes
    n = re.sub(
        r",?\s*(Inc\.?|Co\.?|Company|LLC|Corp\.?)\s*$", "", n, flags=re.IGNORECASE
    ).strip()
    n = re.sub(r",\s*$", "", n).strip()
    return name_map.get(n, n)


# ---------------------------------------------------------------------------
# Cell classification (manual format)
# ---------------------------------------------------------------------------


def is_blank(v):
    if v is None:
        return True
    s = str(v).strip().lower()
    return s in ("", "na", "n/a", "see below")


def starts_x(v):
    if v is None:
        return False
    s = str(v).strip()
    return bool(re.match(r"^[Xx]\s*,", s))


def classify(v):
    """Universal classifier for any status category (manual format).

    Blank / NA / N/A → not tracked (blank)
    X, Received, Complete, or starts with "X," → received
    Anything else → outstanding (O, Required, descriptive text, etc.)
    """
    if is_blank(v):
        return "blank"
    if starts_x(v):
        return "received"
    s = str(v).strip()
    sl = s.lower()
    if sl in ("x", "received", "complete"):
        return "received"
    return "outstanding"


def extract_detail(v):
    """Pull out descriptive text for attic/as-built outstanding items."""
    if is_blank(v):
        return None
    if starts_x(v):
        return None
    s = str(v).strip()
    sl = s.lower()
    if sl in ("x", "received", "o", "complete", "required"):
        return None
    if s.startswith("O("):
        return None
    return s


def filter_note(v):
    """Keep meaningful notes, skip bare status markers."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    sl = s.lower()
    skip = [
        "o", "x", "received", "needed", "o needed", "x needed",
        "see below", "shop dwgs needed", "received needed", "none", "na", "n/a",
    ]
    if sl in skip:
        return None
    if sl.endswith(" needed") and len(s) < 50:
        return None
    if len(s) < 3:
        return None
    return s


# ---------------------------------------------------------------------------
# Note summarization
# ---------------------------------------------------------------------------


def summarize_notes(notes_list):
    """Combine multiple notes into a single summary sentence."""
    if not notes_list:
        return None
    seen = set()
    unique = []
    for n in notes_list:
        key = n.strip().lower()
        if key not in seen:
            seen.add(key)
            unique.append(n.strip())
    if not unique:
        return None
    if len(unique) == 1:
        return unique[0]
    parts = [n.rstrip(".") for n in unique]
    return "; ".join(parts) + "."


# ---------------------------------------------------------------------------
# Shared constants
# ---------------------------------------------------------------------------

CATEGORY_KEYS = ["warranty", "om", "asbuilt", "attic", "training"]
CAT_SHORT = {"warranty": "w", "om": "om", "asbuilt": "ab", "attic": "at", "training": "tr"}
CAT_LABELS = {"warranty": "Warranty", "om": "O&M", "asbuilt": "As-Built", "attic": "Attic Stock", "training": "Training"}


def make_empty_sub(active_cats):
    """Create an empty sub data dict with zeroed counters for active categories."""
    d = {"items": [], "raw_notes": []}
    for c in active_cats:
        short = CAT_SHORT[c]
        d[f"{short}_recv"] = 0
        d[f"{short}_out"] = 0
        d[f"{short}_total"] = 0
    return d


def compute_totals(subs, active_cats):
    """Sum up recv/out/total across all subs."""
    totals = {}
    for cat in active_cats:
        short = CAT_SHORT[cat]
        totals[f"{short}_recv"] = sum(d.get(f"{short}_recv", 0) for d in subs.values())
        totals[f"{short}_out"] = sum(d.get(f"{short}_out", 0) for d in subs.values())
        totals[f"{short}_total"] = sum(d.get(f"{short}_total", 0) for d in subs.values())
    return totals


def finalize_subs(subs):
    """Post-process all subs: summarize notes, clean up."""
    for d in subs.values():
        d["note_summary"] = summarize_notes(d["raw_notes"])
        del d["raw_notes"]


# ---------------------------------------------------------------------------
# Manual format reader
# ---------------------------------------------------------------------------


def read_manual(ws, cols, name_map, training_expected=None):
    """Read manual closeout spreadsheet and return structured sub data."""
    active_cats = [c for c in CATEGORY_KEYS if c in cols]
    subs = defaultdict(lambda: make_empty_sub(active_cats))

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_name = row[cols["sub"]] if cols["sub"] < len(row) else None
        sub = normalize_sub(raw_name, name_map)
        if not sub:
            continue

        spec = str(row[cols["spec"]] if "spec" in cols and cols["spec"] < len(row) else "") or ""
        spec = str(spec).strip() if spec else ""
        desc = str(row[cols["desc"]] if "desc" in cols and cols["desc"] < len(row) else "") or ""
        desc = str(desc).strip() if desc else ""

        item = {"spec": spec, "desc": desc}
        d = subs[sub]

        for cat in active_cats:
            ci = cols[cat]
            v = row[ci] if ci < len(row) else None
            short = CAT_SHORT[cat]
            status = classify(v)
            item[short] = status

            if status != "blank":
                d[f"{short}_total"] += 1
                if status == "received":
                    d[f"{short}_recv"] += 1
                else:
                    d[f"{short}_out"] += 1

            if cat in ("attic", "asbuilt") and status == "outstanding":
                detail = extract_detail(v)
                if detail:
                    item[f"{short}_detail"] = detail

        if "notes" in cols and cols["notes"] < len(row):
            note = filter_note(row[cols["notes"]])
            if note:
                item["note"] = note
                d["raw_notes"].append(note)

        d["items"].append(item)

    finalize_subs(subs)
    totals = compute_totals(subs, active_cats)
    return dict(subs), totals, active_cats


# ---------------------------------------------------------------------------
# Procore format reader
# ---------------------------------------------------------------------------

# Keyword patterns for detecting which closeout categories a title belongs to.
# Order matters — checked in sequence, a title can match multiple categories.
TITLE_CATEGORY_PATTERNS = {
    "warranty": [
        r"\bwarrant(?:y|ies)\b",
    ],
    "om": [
        r"\bo\s*&\s*m\b",
        r"\bo\s+and\s+m\b",
        r"\bmaintenance\b",
        r"\boperation\b",
    ],
    "asbuilt": [
        r"\bas[\s-]?built\b",
        r"\bas[\s-]?build\b",
    ],
    "attic": [
        r"\battic\s*stock\b",
        r"\bspare\s*parts?\b",
    ],
    "training": [
        r"\btraining\b",
        r"\bcommissioning\b",
    ],
}

# Procore Type values that we care about for closeout tracking.
PROCORE_CLOSEOUT_TYPES = {"closeout", "as-built"}


def classify_title_categories(title, procore_type):
    """Determine which closeout categories a Procore row belongs to based on title + type.

    Returns a set of category keys (e.g. {'warranty', 'om', 'attic'}).
    A single row can belong to multiple categories.
    """
    cats = set()
    title_lower = (title or "").lower()
    type_lower = (procore_type or "").lower()

    # If Procore type is explicitly "As-Built", always include asbuilt
    if type_lower == "as-built":
        cats.add("asbuilt")

    # Scan title for category keywords
    for cat, patterns in TITLE_CATEGORY_PATTERNS.items():
        for pat in patterns:
            if re.search(pat, title_lower):
                cats.add(cat)
                break

    # Also scan description / type field for as-built if title mentions it
    if re.search(r"\bas[\s-]?built?\b", title_lower):
        cats.add("asbuilt")

    # If type is "Closeout" but no specific keywords matched,
    # default to O&M since most generic closeout submittals are O&M docs
    if type_lower == "closeout" and not cats:
        cats.add("om")

    return cats


def parse_procore_spec(spec_section_val):
    """Parse Procore's 'Spec Section' column like '23 0000 - Heating, Ventilating...'
    into (spec_number, spec_description)."""
    if not spec_section_val:
        return ("", "")
    s = str(spec_section_val).strip()
    # Format: "23 0000 - Description" or just text
    m = re.match(r"^([\d\s]+)\s*-\s*(.+)$", s)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    # Could be just a number
    if re.match(r"^[\d\s]+$", s):
        return (s.strip(), "")
    return ("", s)


def read_procore(ws, pcols, name_map):
    """Read Procore submittal log export and return structured sub data.

    Only includes rows relevant to closeout tracking:
    - Type = 'Closeout' or 'As-Built'
    - Or title keywords match a closeout category

    Skips types like Product Info, Shop Drawing, Sample, Document, Plans, Complete, Other
    unless their title explicitly contains closeout keywords.
    """
    active_cats = list(CATEGORY_KEYS)  # Procore always has all 5 potential categories
    subs = defaultdict(lambda: make_empty_sub(active_cats))
    skipped = 0
    included = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Extract fields
        procore_type = str(row[pcols["type"]] if "type" in pcols and pcols["type"] < len(row) else "") or ""
        procore_type = procore_type.strip()
        procore_status = str(row[pcols["status"]] if "status" in pcols and pcols["status"] < len(row) else "") or ""
        procore_status = procore_status.strip().lower()
        title = str(row[pcols["title"]] if "title" in pcols and pcols["title"] < len(row) else "") or ""
        title = title.strip()
        raw_sub = str(row[pcols["responsible_contractor"]] if "responsible_contractor" in pcols and pcols["responsible_contractor"] < len(row) else "") or ""
        raw_sub = raw_sub.strip()
        spec_section = row[pcols["spec_section"]] if "spec_section" in pcols and pcols["spec_section"] < len(row) else None

        # Skip rows with no contractor
        if not raw_sub:
            skipped += 1
            continue

        # Skip drafts — they aren't submitted yet
        if procore_status == "draft":
            skipped += 1
            continue

        # Only include rows where Procore Type is Closeout or As-Built.
        # Other types (Shop Drawing, Product Info, etc.) are regular construction submittals,
        # not closeout deliverables, even if their titles contain keywords like "O&M".
        type_lower = procore_type.lower()
        if type_lower not in PROCORE_CLOSEOUT_TYPES:
            skipped += 1
            continue

        # This row is relevant — process it
        included += 1
        sub = normalize_sub(raw_sub, name_map)
        if not sub:
            skipped += 1
            continue

        spec_num, spec_desc = parse_procore_spec(spec_section)
        # Use title as the description if spec section provided a description, otherwise title IS the desc
        desc = title if title else spec_desc

        item = {"spec": spec_num, "desc": desc}
        d = subs[sub]

        # Determine received vs outstanding based on Procore status
        # Closed = received, Open = outstanding
        row_status = "received" if procore_status == "closed" else "outstanding"

        # Use title keywords to determine which of the 5 categories this row belongs to
        cats = classify_title_categories(title, procore_type)

        for cat in active_cats:
            short = CAT_SHORT[cat]
            if cat in cats:
                item[short] = row_status
                d[f"{short}_total"] += 1
                if row_status == "received":
                    d[f"{short}_recv"] += 1
                else:
                    d[f"{short}_out"] += 1
            else:
                item[short] = "blank"

        d["items"].append(item)

    # Remove categories that have zero total items across all subs
    actual_cats = []
    for cat in active_cats:
        short = CAT_SHORT[cat]
        total = sum(d.get(f"{short}_total", 0) for d in subs.values())
        if total > 0:
            actual_cats.append(cat)

    finalize_subs(subs)
    totals = compute_totals(subs, actual_cats)

    print(f"  Procore: {included} closeout rows included, {skipped} non-closeout rows skipped")
    return dict(subs), totals, actual_cats


# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------


def build_html(subs, totals, active_cats, title, primary, secondary, default_on, updated_date):
    """Generate self-contained HTML dashboard string."""

    cat_colors = {
        "warranty": primary,
        "om": secondary,
        "asbuilt": "#2E7D32",
        "attic": "#E65100",
        "training": "#7B1FA2",
    }

    json_subs = {}
    for name, d in subs.items():
        entry = {}
        for cat in active_cats:
            short = CAT_SHORT[cat]
            entry[f"{short}_recv"] = d.get(f"{short}_recv", 0)
            entry[f"{short}_out"] = d.get(f"{short}_out", 0)
            entry[f"{short}_total"] = d.get(f"{short}_total", 0)
        entry["note_summary"] = d.get("note_summary")
        entry["items"] = d["items"]
        json_subs[name] = entry

    data = {"subs": json_subs, "totals": totals}
    data_json = json.dumps(data, ensure_ascii=False)

    toggle_html = ""
    for cat in active_cats:
        short = CAT_SHORT[cat]
        label = CAT_LABELS[cat]
        if cat == "asbuilt":
            label = "As-Built Dwgs"
        active_cls = " active" if cat in default_on else ""
        toggle_html += f'<div class="toggle t{short}{active_cls}" onclick="tog(\'{short}\')">{label}</div>\n'

    act_js = ", ".join(f'{CAT_SHORT[c]}: {"true" if c in default_on else "false"}' for c in active_cats)
    bar_labels = {CAT_SHORT[c]: CAT_LABELS[c] for c in active_cats}
    bar_labels_js = json.dumps(bar_labels)

    toggle_css = ""
    for cat in active_cats:
        short = CAT_SHORT[cat]
        color = cat_colors.get(cat, "#999")
        toggle_css += f".toggle.t{short}{{border-color:{color}}}.toggle.t{short}.active{{background:{color}}}\n"

    fill_css = ""
    for cat in active_cats:
        short = CAT_SHORT[cat]
        color = cat_colors.get(cat, "#999")
        fill_css += f".col-{short} .prog-fill{{background:{color}}}\n"

    detail_css = ""
    for cat in active_cats:
        short = CAT_SHORT[cat]
        color = cat_colors.get(cat, "#999")
        detail_css += f".detail-items.di-{short}{{border-left-color:{color}}}\n"

    th_cols = ""
    for cat in active_cats:
        short = CAT_SHORT[cat]
        label = CAT_LABELS[cat]
        th_cols += f'<th class="col-{short}">{label}</th>\n'

    active_shorts = [CAT_SHORT[c] for c in active_cats]
    active_shorts_js = json.dumps(active_shorts)

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f5f7fa;color:#101820}}
.ctr{{max-width:1400px;margin:0 auto;padding:20px}}
header{{background:{primary};color:#fff;padding:24px 32px;border-radius:12px;margin-bottom:24px}}
header h1{{font-size:24px;font-weight:700}}
header p{{opacity:.8;margin-top:4px;font-size:14px}}
.kpi-row{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:24px}}
.kpi{{background:#fff;border-radius:10px;padding:16px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
.kpi .val{{font-size:28px;font-weight:700;color:{primary}}}
.kpi .lbl{{font-size:12px;color:#7C878E;margin-top:4px;text-transform:uppercase;letter-spacing:.5px}}
.controls{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:24px;align-items:center}}
.toggle-group{{display:flex;gap:8px;flex-wrap:wrap}}
.toggle{{padding:8px 16px;border-radius:8px;border:2px solid #ddd;background:#fff;cursor:pointer;font-size:13px;font-weight:600;transition:all .2s;user-select:none}}
.toggle.active{{color:#fff}}
{toggle_css}
.search-box{{padding:8px 14px;border:2px solid #ddd;border-radius:8px;font-size:14px;min-width:200px}}
select{{padding:8px 14px;border:2px solid #ddd;border-radius:8px;font-size:14px;background:#fff}}
.charts-row{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:24px}}
@media(max-width:900px){{.charts-row{{grid-template-columns:1fr}}}}
.chart-card{{background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
.chart-card h3{{font-size:15px;color:{primary};margin-bottom:12px}}
.sub-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:16px;margin-bottom:24px}}
.sub-card{{background:#fff;border-radius:10px;padding:18px;box-shadow:0 1px 3px rgba(0,0,0,.08);border-left:5px solid #ccc}}
.sub-card h4{{font-size:16px;color:{primary};margin-bottom:10px}}
.prog-row{{display:flex;align-items:center;gap:8px;margin-bottom:6px;font-size:13px}}
.prog-row .lbl{{min-width:70px;font-weight:600}}
.prog-bar{{flex:1;height:10px;background:#e9ecef;border-radius:5px;overflow:hidden}}
.prog-fill{{height:100%;border-radius:5px;transition:width .3s}}
.prog-row .pct{{min-width:45px;text-align:right;font-weight:600}}
{fill_css}
.detail-items{{font-size:12px;color:#555;margin-top:4px;margin-bottom:6px;padding-left:8px;border-left:2px solid #ccc;display:none}}
.detail-items.vis{{display:block}}
{detail_css}
.detail-items div{{margin-bottom:3px}}
.details-toggle{{display:inline-flex;align-items:center;gap:6px;padding:6px 14px;border-radius:8px;border:2px solid #ddd;background:#fff;cursor:pointer;font-size:13px;font-weight:600;margin-left:12px;transition:all .2s;user-select:none;vertical-align:middle}}
.details-toggle.active{{border-color:{primary};background:{primary};color:#fff}}
.note{{font-size:12px;color:#7C878E;margin-top:8px;padding:8px;background:#f8f9fa;border-radius:6px;font-style:italic}}
.detail-table{{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
.detail-table th{{background:{primary};color:#fff;padding:10px 12px;text-align:left;font-size:13px;font-weight:600}}
.detail-table td{{padding:8px 12px;border-bottom:1px solid #eee;font-size:13px}}
.detail-table tr:hover{{background:#f5f7fa}}
.recv{{color:#43A047;font-weight:600}}.out{{color:#E53935;font-weight:600}}.blank{{color:#ccc}}
h2{{font-size:18px;color:{primary};margin-bottom:16px;margin-top:8px}}
.section{{margin-bottom:24px}}
</style>
</head>
<body>
<div class="ctr">
<header>
<h1>{title}</h1>
<p>Subcontractor closeout tracking &bull; Last updated {updated_date}</p>
</header>

<div class="kpi-row" id="kpis"></div>

<div class="controls">
<div class="toggle-group">
{toggle_html}
</div>
<input class="search-box" type="text" placeholder="Search subs..." oninput="doSearch(this.value)">
<select id="subFilter" onchange="doFilter(this.value)"><option value="">All Subcontractors</option></select>
</div>

<div class="charts-row">
<div class="chart-card"><h3>Overall Progress by Category</h3><canvas id="chart1"></canvas></div>
<div class="chart-card"><h3>Outstanding Items by Subcontractor</h3><canvas id="chart2"></canvas></div>
</div>

<div class="section">
<div style="display:flex;align-items:center;gap:12px;margin-bottom:16px"><h2 style="margin:0">Subcontractor Summaries</h2><div class="details-toggle" onclick="togDetails()">Show Items</div></div>
<div class="sub-grid" id="subGrid"></div>
</div>

<div class="section">
<h2>Full Detail</h2>
<div style="overflow-x:auto">
<table class="detail-table" id="detailTable">
<thead><tr>
<th>Subcontractor</th>
<th>Spec</th>
<th>Description</th>
{th_cols}
<th>Notes</th>
</tr></thead>
<tbody id="detailBody"></tbody>
</table>
</div>
</div>
</div>

<script>
const D={data_json};
const S=D.subs,T=D.totals;
const catKeys={active_shorts_js};
const barLabels={bar_labels_js};
let act={{{act_js}}};
let showDetails=false;
let searchQ='',filterSub='';

function tog(t){{
act[t]=!act[t];
document.querySelector('.t'+t).classList.toggle('active');
rAll();
}}

function doSearch(v){{searchQ=v.toLowerCase();rCards();rTable();}}
function doFilter(v){{filterSub=v;rCards();rTable();}}

function matchSub(name){{
if(filterSub&&name!==filterSub)return false;
if(searchQ&&!name.toLowerCase().includes(searchQ))return false;
return true;
}}

function rKPI(){{
let totOut=0,totAll=0;
for(let[n,d]of Object.entries(S)){{
catKeys.forEach(k=>{{
if(act[k]){{totOut+=d[k+'_out']||0;totAll+=d[k+'_total']||0}}
}});
}}
let totRecv=totAll-totOut;
let pct=totAll?Math.round(totRecv/totAll*100):100;
let subsComplete=Object.entries(S).filter(([n,d])=>{{
let o=0,t2=0;
catKeys.forEach(k=>{{if(act[k]){{o+=d[k+'_out']||0;t2+=d[k+'_total']||0}}}});
return t2>0&&o===0;
}}).length;
document.getElementById('kpis').innerHTML=`
<div class="kpi"><div class="val">${{pct}}%</div><div class="lbl">Overall Complete</div></div>
<div class="kpi"><div class="val">${{totRecv}}</div><div class="lbl">Items Received</div></div>
<div class="kpi"><div class="val">${{totOut}}</div><div class="lbl">Outstanding</div></div>
<div class="kpi"><div class="val">${{totAll}}</div><div class="lbl">Total Tracked</div></div>
<div class="kpi"><div class="val">${{subsComplete}}/${{Object.keys(S).length}}</div><div class="lbl">Subs Complete</div></div>
`;
}}

function rCharts(){{
if(window.c1)window.c1.destroy();
if(window.c2)window.c2.destroy();
let cats=[],recv=[],out=[];
catKeys.forEach(k=>{{
if(act[k]){{cats.push(barLabels[k]);recv.push(T[k+'_recv']||0);out.push(T[k+'_out']||0)}}
}});
window.c1=new Chart(document.getElementById('chart1'),{{
type:'bar',data:{{labels:cats,datasets:[
{{label:'Received',data:recv,backgroundColor:'#43A047'}},
{{label:'Outstanding',data:out,backgroundColor:'#E53935'}}
]}},options:{{responsive:true,plugins:{{legend:{{position:'bottom'}}}},scales:{{x:{{stacked:true}},y:{{stacked:true,beginAtZero:true}}}}}}
}});
let sn=[],sv=[];
for(let[n,d]of Object.entries(S)){{
let o=0;
catKeys.forEach(k=>{{if(act[k])o+=d[k+'_out']||0}});
if(o>0){{sn.push(n);sv.push(o)}}
}}
let idx=[...sv.keys()].sort((a,b)=>sv[b]-sv[a]);
sn=idx.map(i=>sn[i]);sv=idx.map(i=>sv[i]);
window.c2=new Chart(document.getElementById('chart2'),{{
type:'bar',data:{{labels:sn,datasets:[{{label:'Outstanding',data:sv,backgroundColor:'{primary}'}}]}},
options:{{indexAxis:'y',responsive:true,plugins:{{legend:{{display:false}}}},scales:{{x:{{beginAtZero:true}}}}}}
}});
}}

function pBar(recv,total,cls){{
let pct=total?Math.round(recv/total*100):0;
return `<div class="prog-row ${{cls}}"><span class="lbl">${{barLabels[cls.replace('col-','')]||cls}}</span><div class="prog-bar"><div class="prog-fill" style="width:${{pct}}%"></div></div><span class="pct">${{recv}}/${{total}}</span></div>`;
}}

function detailList(items,cat,cls){{
let outs=items.filter(i=>i[cat]==='outstanding'&&(i.spec||i.desc));
if(!outs.length)return '';
let vis=showDetails?' vis':'';
let html=`<div class="detail-items di-${{cls}}${{vis}}">`;
outs.forEach(i=>{{
let label=i.spec?(i.spec+' '+i.desc).trim():i.desc;
let extra='';
if(i[cat+'_detail'])extra=' — '+i[cat+'_detail'];
html+=`<div>• ${{label}}${{extra}}</div>`;
}});
return html+'</div>';
}}

function rSC(name,d){{
let outTotal=0,tracked=0;
let bars='';
catKeys.forEach(k=>{{
if(act[k]&&(d[k+'_total']||0)>0){{
bars+=pBar(d[k+'_recv']||0,d[k+'_total'],`col-${{k}}`);
if((d[k+'_out']||0)>0)bars+=detailList(d.items,k,k);
outTotal+=d[k+'_out']||0;
tracked+=d[k+'_total']||0;
}}
}});
if(!bars)return '';
let pct=tracked?Math.round((tracked-outTotal)/tracked*100):100;
let hue=Math.round(pct*1.2);
let sat=pct===100?'60%':'75%';
let lgt=pct>60?'38%':pct>30?'45%':'42%';
let bColor=`hsl(${{hue}},${{sat}},${{lgt}})`;
let noteHtml='';
if(d.note_summary)noteHtml=`<div class="note">Note: ${{d.note_summary}}</div>`;
return `<div class="sub-card" style="border-left-color:${{bColor}}"><h4>${{name}}</h4>${{bars}}${{noteHtml}}</div>`;
}}

function rCards(){{
let html='';
for(let[n,d]of Object.entries(S)){{
if(!matchSub(n))continue;
let c=rSC(n,d);
if(c)html+=c;
}}
document.getElementById('subGrid').innerHTML=html||'<p style="color:#7C878E">No matching subs</p>';
}}

function stClass(v){{return v==='received'?'recv':v==='outstanding'?'out':'blank'}}
function stLabel(v){{return v==='received'?'✓':v==='outstanding'?'✗':''}}

function rTable(){{
let html='';
for(let[n,d]of Object.entries(S)){{
if(!matchSub(n))continue;
d.items.forEach(item=>{{
let show=false;
catKeys.forEach(k=>{{if(act[k]&&item[k]!=='blank')show=true}});
if(!show)return;
html+=`<tr><td>${{n}}</td><td>${{item.spec}}</td><td>${{item.desc}}</td>`;
catKeys.forEach(k=>{{
let v=item[k]||'blank';
let extra='';
if(v==='outstanding'&&item[k+'_detail'])extra=item[k+'_detail'];
html+=`<td class="col-${{k}} ${{stClass(v)}}">${{extra||stLabel(v)}}</td>`;
}});
html+=`<td>${{item.note||''}}</td></tr>`;
}});
}}
document.getElementById('detailBody').innerHTML=html;
catKeys.forEach(t=>{{
document.querySelectorAll('.col-'+t).forEach(e=>e.style.display=act[t]?'':'none');
}});
}}

function togDetails(){{
showDetails=!showDetails;
document.querySelector('.details-toggle').classList.toggle('active');
document.querySelector('.details-toggle').textContent=showDetails?'Hide Items':'Show Items';
document.querySelectorAll('.detail-items').forEach(e=>{{
e.classList.toggle('vis',showDetails);
}});
}}

function rAll(){{rKPI();rCharts();rCards();rTable();}}

let sel=document.getElementById('subFilter');
Object.keys(S).sort().forEach(n=>{{
let o=document.createElement('option');o.value=n;o.textContent=n;sel.appendChild(o);
}});

rAll();
</script>
</body>
</html>'''

    return html


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="Build closeout dashboard from Excel")
    parser.add_argument("--input", required=True, help="Path to Excel file (manual or Procore export)")
    parser.add_argument("--output", required=True, help="Path for output HTML")
    parser.add_argument("--title", default="Closeout Dashboard", help="Dashboard title")
    parser.add_argument("--primary", default="#174A5B", help="Primary brand color hex")
    parser.add_argument("--secondary", default="#5489A3", help="Secondary brand color hex")
    parser.add_argument("--default-on", default="warranty,om",
                        help="Comma-separated categories on by default (warranty,om,asbuilt,attic,training)")
    parser.add_argument("--name-map", default=None, help="JSON file with sub name mappings")
    parser.add_argument("--training-expected", default=None,
                        help="Comma-separated sub names expected to have training")
    parser.add_argument("--format", default="auto", choices=["auto", "manual", "procore"],
                        help="Input format: auto-detect (default), manual, or procore")

    args = parser.parse_args()

    # Parse default-on
    default_on_raw = [x.strip().lower() for x in args.default_on.split(",")]
    default_on_map = {"warranty": "warranty", "om": "om", "asbuilt": "asbuilt",
                      "attic": "attic", "training": "training",
                      "as-built": "asbuilt", "as_built": "asbuilt"}
    default_on = set()
    for item in default_on_raw:
        if item in default_on_map:
            default_on.add(default_on_map[item])

    training_expected = None
    if args.training_expected:
        training_expected = [x.strip() for x in args.training_expected.split(",")]

    name_map = load_name_map(args.name_map)

    print(f"Reading: {args.input}")
    wb = load_workbook(args.input, data_only=True)
    ws = wb.active

    # Detect or use specified format
    if args.format == "auto":
        fmt = detect_format(ws)
    else:
        fmt = args.format
    print(f"Format: {fmt}")

    if fmt == "procore":
        pcols = detect_procore_columns(ws)
        print(f"Procore columns: {pcols}")
        subs, totals, active_cats = read_procore(ws, pcols, name_map)
    else:
        cols = detect_manual_columns(ws)
        print(f"Manual columns: {cols}")
        subs, totals, active_cats = read_manual(ws, cols, name_map, training_expected)

    print(f"Found {len(subs)} subcontractors across {len(active_cats)} categories")
    for cat in active_cats:
        short = CAT_SHORT[cat]
        recv = totals.get(f"{short}_recv", 0)
        out = totals.get(f"{short}_out", 0)
        total = totals.get(f"{short}_total", 0)
        print(f"  {CAT_LABELS[cat]}: {recv}/{total} received, {out} outstanding")

    updated_date = date.today().strftime("%B %-d, %Y")
    html = build_html(subs, totals, active_cats, args.title, args.primary,
                      args.secondary, default_on, updated_date)

    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Dashboard written to: {args.output}")


if __name__ == "__main__":
    main()
